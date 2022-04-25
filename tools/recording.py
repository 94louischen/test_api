# -*- coding: utf-8 -*-
# @Time : 2021/9/7 14:38
# @Author : chenxuan
"""
@Desc  : 录制接口，生成用例文件
基于mitmproxy实现，会包含css/html/png等后缀链接
参考资料：
https://blog.wolfogre.com/posts/usage-of-mitmproxy/
https://www.cnblogs.com/liuwanqiu/p/10697373.html
"""
import datetime

import mitmproxy.http
import openpyxl
from mitmproxy import ctx

import json


class RecordCases:

    def __init__(self, filter_url: str, filename: str = "data/infra_cases_test.xlsx", sheet_name: str = "录制用例数据"):
        self.filter_url = filter_url
        self.excel_row = [
            'case_id',
            'module_name',
            'title',
            'url',
            'headers',
            'data',
            'method',
            'expected',
            'actual',
            'result',
            'check_sql',
            'Extraction',
            'enable',
            'create_time',
            'update_time',
            'auther',
            'data_type']
        self.cases = [self.excel_row]
        self.counter = 1
        self.file = filename
        self.sheet_name = sheet_name

    def response(self, flow: mitmproxy.http.HTTPFlow):
        """
        mitmproxy抓包处理响应，在这里汇总需要数据
        :param flow:
        :return:
        """
        url = flow.request.url
        if url.find('.css') == -1 and url.find('.js') == -1 and url.find(self.filter_url) != -1:
            # 编号
            case_id = str(self.counter)
            # 标题
            title = "mitmproxy录制接口" + case_id
            data = flow.request.text
            method = flow.request.method.lower()
            url = flow.request.url
            Referer = flow.request.headers["Host"]
            try:
                content_type = flow.request.headers['Content-Type']
            except KeyError:
                content_type = ''
            if 'multipart' in content_type:
                data_type = "file"
            elif 'form' in content_type:
                data_type = "data"
            elif 'json' in content_type:
                data_type = 'json'
            else:
                data_type = 'params'
                if '?' in url:
                    data = url.split('?')[1]
            headers = json.dumps({"Content-Type": content_type, "Referer": Referer})
            data = self.handle_form(data)
            # 预期结果
            try:
                expect = json.dumps(
                    {".": json.loads(flow.response.text)}, ensure_ascii=False)
            except Exception as e:
                ctx.log.error(e)
                expect = '{}'
            # 日志
            # ctx.log.info(case_id)
            # ctx.log.info(title)
            # ctx.log.info(url)
            # ctx.log.info(headers)
            # ctx.log.info(method)
            # ctx.log.info(data)
            ctx.log.info(flow.response.text)
            case = [
                case_id,
                None,
                title,
                url,
                headers,
                data,
                method,
                None,
                None,
                None,
                None,
                None,
                1,
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '陈旋',
                data_type]
            self.cases.append(case)
            self.counter += 1
            self.excel_cases()

    def handle_form(self, data: str):
        """
        处理 Content-Type:	application/x-www-form-urlencoded
        默认生成的数据 username=admin&password=123456
        :param data: 获取的data 类似这样  username=admin&password=123456
        :return:
        """
        data_dict = {}
        if data.startswith('{') and data.endswith('}'):
            return data
        try:
            for i in data.split('&'):
                data_dict[i.split('=')[0]] = i.split('=')[1]
            return json.dumps(data_dict)
        except IndexError:
            return ''

    def excel_cases(self):
        """
        对二维列表cases进行循环并将内容写入单元格中
        :return:
        """
        workbook = openpyxl.load_workbook(self.file)
        try:
            worksheet = workbook[self.sheet_name]
        except KeyError as e:
            ctx.log.error(e)
            worksheet = workbook.create_sheet('录制用例数据', 0)
        for x in range(len(self.cases)):
            for y in range(len(self.cases[x])):
                worksheet.cell(x + 1, y + 1).value = self.cases[x][y]
        workbook.save(self.file)
        workbook.close()


addons = [RecordCases('test.xxxx.com')]  # 不同的项目组修改对应的域名
