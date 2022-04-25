# -*- coding=utf-8 -*-
import openpyxl


class Case:

    def __init__(self):
        self.case_id = None
        self.module_name = None
        self.title = None
        self.url = None
        self.headers = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None
        self.extraction = None
        self.enable = None
        self.create_time = None
        self.update_time = None
        self.auther = None
        self.data_type = None
        self.pro_runing = None


class DoExcel:

    def __init__(self, files, sheet_name):
        self.files = files
        self.sheet_name = sheet_name
        self.lw = openpyxl.load_workbook(files)
        self.sheet = self.lw[sheet_name]

    def read_excel(self):
        cases = []
        for row in range(2, self.sheet.max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row, 1).value
            case.module_name = self.sheet.cell(row, 2).value
            case.title = self.sheet.cell(row, 3).value
            case.url = self.sheet.cell(row, 4).value
            case.headers = self.sheet.cell(row, 5).value
            case.data = self.sheet.cell(row, 6).value
            case.method = self.sheet.cell(row, 7).value
            case.expected = self.sheet.cell(row, 8).value
            case.actual = self.sheet.cell(row, 9).value
            case.result = self.sheet.cell(row, 10).value
            case.check_sql = self.sheet.cell(row, 11).value
            case.extraction = self.sheet.cell(row, 12).value
            case.enable = self.sheet.cell(row, 13).value
            case.create_time = self.sheet.cell(row, 14).value
            case.update_time = self.sheet.cell(row, 15).value
            case.auther = self.sheet.cell(row, 16).value
            case.data_type = self.sheet.cell(row, 17).value
            case.pro_runing = self.sheet.cell(row, 18).value
            cases.append(case)
        self.lw.close()
        return cases

    def write_excel(self, case_id, actual, result):
        sheet = self.lw[self.sheet_name]
        sheet.cell(case_id + 1, 8).value = actual
        sheet.cell(case_id + 1, 9).value = result
        self.lw.save(self.files)
        self.lw.close()


if __name__ == '__main__':
    pass
